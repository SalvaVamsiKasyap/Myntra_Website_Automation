import openpyxl

book = openpyxl.load_workbook(
    r"C:\Users\Lenovo\Downloads\pypxl.xlsx",keep_vba=True)

sheet = book.active

cell = sheet.cell(row=1, column=2)

print(cell.value)

sheet.cell(row=2, column=2).value = 'kasyap'

book.save(r"C:\Users\Lenovo\Downloads\pypxl.xlsx")

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['B2'].value)

#book = openpyxl.open(r"C:\Users\Lenovo\Downloads\Verification.xlsx")
